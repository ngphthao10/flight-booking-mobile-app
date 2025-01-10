from flask import Blueprint, jsonify, request, current_app, send_file
from app.decorators import admin_required
import requests
from datetime import datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

report = Blueprint('report', __name__)

def get_report_title(status, time_filter):
    titles = {
        'today': 'HÔM NAY',
        'yesterday': 'HÔM QUA',
        'last_7_days': '7 NGÀY QUA',
        'last_14_days': '14 NGÀY QUA', 
        'last_21_days': '21 NGÀY QUA',
        'last_month': '1 THÁNG QUA',
        'last_3_month': '3 THÁNG QUA'
    }
    return f"{status} {titles.get(time_filter, '')}"

@report.route('/export-market-share')
@admin_required
def export_market_share():
    wb = None
    temp_file = None
    try:
        time_filter = request.args.get('time', 'today')
        
        api_url = f"{current_app.config['API_URL']}/api/report/market_share"
        response = requests.get(api_url, params={'time': time_filter})
        response.raise_for_status()
        
        data = response.json()
        
        if data['status'] != 'success':
            return jsonify({"error": "Failed to fetch data from API"}), 400
            
        # Tạo workbook và thiết lập styles
        wb = Workbook()
        ws = wb.active
        
        # Định nghĩa styles với font Times New Roman
        header_font = Font(name='Times New Roman', bold=True, size=13)
        title_font = Font(name='Times New Roman', bold=True, size=14)
        data_font = Font(name='Times New Roman', size=13)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Thiết lập chiều cao hàng mặc định
        ws.sheet_properties.customHeight = True
        ws.row_dimensions[1].height = 30
        
        # Thêm tiêu đề báo cáo
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = get_report_title('BÁO CÁO THỊ PHẦN HÃNG HÀNG KHÔNG',time_filter)
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        
        # Thêm thông tin thời gian và người xuất
        ws['A2'] = f"Thời gian xuất: {datetime.now().strftime('%H:%M:%S %d/%m/%Y')}"
        ws['A3'] = "Người xuất: admin"
        for row in [2, 3]:
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
        
        # Thêm header bảng
        headers = ['Mã hàng', 'Tên hàng', 'Số lượng đặt chỗ', 'Thị phần (%)']
        alignments = [left_align, left_align, right_align, right_align]
        
        for col, (header, alignment) in enumerate(zip(headers, alignments), 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment
            cell.fill = header_fill
        
        # Thêm dữ liệu
        row_num = 5
        total_bookings = 0
        
        for item in data['data']:
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if row_num % 2 == 0 else None
            
            cells = [
                (item['airline_code'], left_align),
                (item['airline_name'], left_align),
                (item['total_bookings'], right_align),
                (f"{round(item['market_share'], 1)}%", right_align)  # Thêm ký hiệu % vào thị phần
            ]
            
            for col, (value, alignment) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = alignment
                if row_fill:
                    cell.fill = row_fill
                
            total_bookings += item['total_bookings']
            row_num += 1
            
        # Thêm dòng tổng cộng
        total_row = row_num
        total_cells = [
            ('Tổng cộng', left_align),
            ('', center_align),
            (total_bookings, right_align),
            ('100%', right_align)  # Thêm ký hiệu % vào tổng thị phần
        ]
        
        for col, (value, alignment) in enumerate(total_cells, 1):
            cell = ws.cell(row=total_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment
            cell.fill = total_fill
            
        # Căn chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Tạo temporary file và lưu
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_file_path)
        
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=f"market_share_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    finally:
        if wb:
            wb.close()
        if temp_file:
            try:
                import threading
                def delete_file():
                    import time
                    time.sleep(3)
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                threading.Thread(target=delete_file).start()
            except:
                pass

@report.route('/export-revenue')
@admin_required
def export_revenue():
    wb = None
    temp_file = None
    try:
        report_type = request.args.get('type', 'monthly')
        month = request.args.get('month', str(datetime.now().month))
        year = request.args.get('year', str(datetime.now().year))
        include_growth = request.args.get('include_growth', 'false')
        
        api_url = f"{current_app.config['API_URL']}/api/report/revenue"
        response = requests.get(api_url, params={
            'type': report_type,
            'month': month,
            'year': year,
            'include_growth': include_growth
        })
        response.raise_for_status()
        
        data = response.json()
        
        if data['status'] != 'success':
            return jsonify({"error": "Failed to fetch data from API"}), 400
            
        wb = Workbook()
        ws = wb.active
        
        header_font = Font(name='Times New Roman', bold=True, size=13)
        title_font = Font(name='Times New Roman', bold=True, size=14)
        data_font = Font(name='Times New Roman', size=13)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        ws.sheet_properties.customHeight = True
        ws.row_dimensions[1].height = 30
        
        title = "BÁO CÁO DOANH THU "
        if report_type == 'monthly':
            if month != 'all':
                title += f"THÁNG {month}"
            if year != 'all':
                title += f" NĂM {year}"
        elif report_type == 'weekly':
            title += "THEO TUẦN"
            if year != 'all':
                title += f" NĂM {year}"
        
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        
        # Thêm thông tin thời gian và người xuất
        ws['A2'] = f"Thời gian xuất: {datetime.now().strftime('%H:%M:%S %d/%m/%Y')}"
        ws['A3'] = "Người xuất: admin"
        for row in [2, 3]:
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
        
        # Thêm header bảng
        if report_type == 'monthly':
            headers = ['Ngày', 'Số đơn hàng', 'Doanh thu (VNĐ)']
        else:
            headers = ['Tuần', 'Năm', 'Số đơn hàng', 'Doanh thu (VNĐ)']
            if include_growth == 'true':
                headers.append('Tăng trưởng (%)')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = header_fill
        
        # Thêm dữ liệu
        row_num = 5
        total_orders = 0
        total_revenue = 0
        
        for item in data['data']:
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if row_num % 2 == 0 else None
            
            if report_type == 'monthly':
                cells = [
                    (datetime.strptime(item['date'], '%Y-%m-%d').strftime('%d/%m/%Y'), center_align),
                    (item['total_orders'], right_align),
                    (item['total_revenue'], right_align)
                ]
            else:
                cells = [
                    (f"Tuần {item['week_number']}", center_align),
                    (item['year'], center_align),
                    (item['total_orders'], right_align),
                    (item['total_revenue'], right_align)
                ]
                if include_growth == 'true':
                    cells.append((f"{item.get('growth_rate', 0)}%", right_align))
            
            for col, (value, alignment) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = alignment
                if row_fill:
                    cell.fill = row_fill
                
                # Format số cho cột doanh thu
                if col == (3 if report_type == 'monthly' else 4):
                    cell.number_format = '#,##0'
            
            total_orders += item['total_orders']
            total_revenue += item['total_revenue']
            row_num += 1
            
        # Thêm dòng tổng cộng
        total_row = row_num
        if report_type == 'monthly':
            total_cells = [
                ('Tổng cộng', center_align),
                (total_orders, right_align),
                (total_revenue, right_align)
            ]
        else:
            total_cells = [
                ('Tổng cộng', center_align),
                ('', center_align),
                (total_orders, right_align),
                (total_revenue, right_align)
            ]
            if include_growth == 'true':
                total_cells.append(('', center_align))
        
        for col, (value, alignment) in enumerate(total_cells, 1):
            cell = ws.cell(row=total_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment
            cell.fill = total_fill
            
            # Format số cho cột doanh thu trong dòng tổng
            if col == (3 if report_type == 'monthly' else 4):
                cell.number_format = '#,##0'
        
        # Căn chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        if report_type == 'weekly':
            ws.column_dimensions['D'].width = 20
            if include_growth == 'true':
                ws.column_dimensions['E'].width = 15
        
        # Tạo temporary file và lưu
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_file_path)
        
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=f"revenue_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    finally:
        if wb:
            wb.close()
        if temp_file:
            try:
                import threading
                def delete_file():
                    import time
                    time.sleep(3)
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                threading.Thread(target=delete_file).start()
            except:
                pass

@report.route('/export-booking-stats')
@admin_required
def export_booking_stats():
    wb = None
    temp_file = None
    try:
        time_range = request.args.get('time_range', 'last_7_days')
        
        api_url = f"{current_app.config['API_URL']}/api/report/booking_stats"
        response = requests.get(api_url, params={
            'time_range': time_range
        })
        response.raise_for_status()
        
        data = response.json()
        
        if data['status'] != 'success':
            return jsonify({"error": "Failed to fetch data from API"}), 400
            
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        header_font = Font(name='Times New Roman', bold=True, size=13)
        title_font = Font(name='Times New Roman', bold=True, size=14)
        data_font = Font(name='Times New Roman', size=13)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Set row height
        ws.sheet_properties.customHeight = True
        ws.row_dimensions[1].height = 30
        
        # Set title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = get_report_title('BÁO CÁO ĐẶT CHỖ', time_range)
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        
        # Add export info
        ws['A2'] = f"Thời gian xuất: {datetime.now().strftime('%H:%M:%S %d/%m/%Y')}"
        ws['A3'] = "Người xuất: admin"
        for row in [2, 3]:
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
        
        # Add headers
        headers = ['Ngày', 'Số lượng đặt chỗ', 'Tổng số hành khách']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = header_fill
        
        # Add data
        row_num = 5
        stats = data['data']['stats']
        
        for item in stats:
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if row_num % 2 == 0 else None
            
            cells = [
                (datetime.strptime(item['date'], '%Y-%m-%d').strftime('%d/%m/%Y'), center_align),
                (item['total_bookings'], right_align),
                (item['total_passengers'], right_align)
            ]
            
            for col, (value, alignment) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = alignment
                if row_fill:
                    cell.fill = row_fill
            
            row_num += 1
        
        # Add totals
        total_row = row_num
        total_cells = [
            ('Tổng cộng', center_align),
            (data['data']['total_bookings'], right_align),
            (data['data']['total_passengers'], right_align)
        ]
        
        for col, (value, alignment) in enumerate(total_cells, 1):
            cell = ws.cell(row=total_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment
            cell.fill = total_fill
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Create and save temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_file_path)
        
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=f"booking_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    finally:
        if wb:
            wb.close()
        if temp_file:
            try:
                import threading
                def delete_file():
                    import time
                    time.sleep(3)
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                threading.Thread(target=delete_file).start()
            except:
                pass

@report.route('/export-baggage-stats')
@admin_required
def export_baggage_stats():
    wb = None
    temp_file = None
    try:
        time_range = request.args.get('time_range', 'last_7_days')
        
        api_url = f"{current_app.config['API_URL']}/api/report/baggage_service_stats"
        response = requests.get(api_url, params={
            'time_range': time_range
        })
        response.raise_for_status()
        
        data = response.json()
        
        if data['status'] != 'success':
            return jsonify({"error": "Failed to fetch data from API"}), 400
            
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        header_font = Font(name='Times New Roman', bold=True, size=13)
        title_font = Font(name='Times New Roman', bold=True, size=14)
        data_font = Font(name='Times New Roman', size=13)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Set row height
        ws.sheet_properties.customHeight = True
        ws.row_dimensions[1].height = 30
        
        # Set title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = get_report_title('BÁO CÁO DỊCH VỤ HÀNH LÝ', time_range)
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        
        # Add export info
        ws['A2'] = f"Thời gian xuất: {datetime.now().strftime('%H:%M:%S %d/%m/%Y')}"
        ws['A3'] = "Người xuất: admin"
        for row in [2, 3]:
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
        
        # Add headers
        headers = ['Trọng lượng', 'Số lượng đặt', 'Doanh thu (VNĐ)', 'Tỷ lệ (%)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.fill = header_fill
        
        # Add data
        row_num = 5
        stats = data['data']['stats']
        
        for item in stats:
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") if row_num % 2 == 0 else None
            
            cells = [
                (item['weight'], center_align),
                (item['bookings'], right_align),
                (item['revenue'], right_align),
                (f"{item['percentage']}%", right_align)
            ]
            
            for col, (value, alignment) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = alignment
                if row_fill:
                    cell.fill = row_fill
                
                # Format số cho cột doanh thu
                if col == 3:
                    cell.number_format = '#,##0'
            
            row_num += 1
        
        # Add totals
        total_row = row_num
        total_cells = [
            ('Tổng cộng', center_align),
            (data['data']['total_bookings'], right_align),
            (data['data']['total_revenue'], right_align),
            ('100%', right_align)
        ]
        
        for col, (value, alignment) in enumerate(total_cells, 1):
            cell = ws.cell(row=total_row, column=col)
            cell.value = value
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = alignment
            cell.fill = total_fill
            
            # Format số cho cột doanh thu trong dòng tổng
            if col == 3:
                cell.number_format = '#,##0'
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Create and save temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_file_path)
        
        return send_file(
            temp_file_path,
            as_attachment=True,
            download_name=f"baggage_service_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
    finally:
        if wb:
            wb.close()
        if temp_file:
            try:
                import threading
                def delete_file():
                    import time
                    time.sleep(3)
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass
                threading.Thread(target=delete_file).start()
            except:
                pass